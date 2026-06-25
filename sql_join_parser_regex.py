"""
Production-Grade SQL Join Parser
=================================
Parses database query logs from a CSV file where:
  - Each NEW record begins on a new line starting with a username token
  - SQL text is multi-line (spans multiple lines within a record)
  - Blank lines appear both between records AND within multi-line SQL fields
  - Some SQL fields are optionally quoted, others are not

Outputs a detailed Excel report of table-to-table join relationships,
aggregated by [left_table, right_table, date] grain.

Usage:
    python sql_join_parser.py --input <path_to_input.csv> --output <path_to_output.xlsx>

Dependencies:
    pip install pandas openpyxl
"""

import re
import sys
import logging
import argparse

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# LOGGING CONFIGURATION
# ---------------------------------------------------------------------------
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[logging.StreamHandler(sys.stdout)],
)
logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# CONSTANTS
# ---------------------------------------------------------------------------

EXPECTED_COLUMNS = 7

# Known service-account prefixes for unique_app counting
SERVICE_PREFIXES = ("SVP", "SVR", "SVC", "SVT", "ETL", "DBA", "OP")

# ---------------------------------------------------------------------------
# COMPILED REGEXES
# ---------------------------------------------------------------------------

# Detects the START of a new record: a line beginning with a compact username token
# (alphanumeric, no spaces) followed by a comma.
RE_RECORD_START = re.compile(r'^([A-Za-z0-9]+)\s*,')

# Date separator pattern used to locate the LogDate field boundary inside a raw record.
# Matches a comma followed by a date token like "3/24/2026" or "5/13/2026 9:49"
RE_DATE_SEP = re.compile(r',\s*(\d{1,2}/\d{1,2}/\d{4}(?:\s+\d{1,2}:\d{2})?)')

# SQL comment stripping
RE_BLOCK_COMMENT = re.compile(r'/\*.*?\*/', re.DOTALL)
RE_LINE_COMMENT  = re.compile(r'--[^\n]*')
RE_WHITESPACE    = re.compile(r'\s+')

# SQL keywords to exclude from table-name resolution
SQL_KEYWORDS = {
    'WHERE', 'ON', 'SET', 'GROUP', 'ORDER', 'HAVING', 'LIMIT', 'UNION',
    'EXCEPT', 'INTERSECT', 'SELECT', 'WITH', 'VALUES', 'INNER', 'LEFT',
    'RIGHT', 'FULL', 'OUTER', 'CROSS', 'NATURAL', 'JOIN', 'FROM', 'AND',
    'OR', 'NOT', 'IN', 'AS', 'BY', 'BETWEEN', 'CASE', 'WHEN', 'THEN',
    'END', 'OVER', 'PARTITION', 'QUALIFY', 'DISTINCT', 'NULL', 'IS',
}

# Captures explicit JOIN clauses and the table they reference
# Handles: INNER JOIN, LEFT JOIN, LEFT OUTER JOIN, RIGHT JOIN, FULL OUTER JOIN,
#          CROSS JOIN, NATURAL JOIN, plain JOIN
RE_JOIN_CLAUSE = re.compile(
    r'(?:INNER\s+JOIN|LEFT\s+(?:OUTER\s+)?JOIN|RIGHT\s+(?:OUTER\s+)?JOIN'
    r'|FULL\s+(?:OUTER\s+)?JOIN|CROSS\s+JOIN|NATURAL\s+JOIN|(?<!\w)JOIN)\s+'
    r'((?:\w+\.)*\w+)'                     # table name with optional schema prefix
    r'(?:\s+(?:AS\s+)?(\w+))?',            # optional alias
    re.IGNORECASE,
)

# Captures FROM clause table reference
RE_FROM_CLAUSE = re.compile(
    r'(?<!\w)FROM\s+((?:\w+\.)*\w+)(?:\s+(?:AS\s+)?(\w+))?',
    re.IGNORECASE,
)

# Full table reference (FROM or JOIN) for alias map building
RE_TABLE_REF = re.compile(
    r'(?:FROM|JOIN)\s+((?:\w+\.)*\w+)(?:\s+(?:AS\s+)?(\w+))?',
    re.IGNORECASE,
)

# ---------------------------------------------------------------------------
# HELPER FUNCTIONS
# ---------------------------------------------------------------------------

def strip_prefix(name: str) -> str:
    """Remove schema/db prefixes: 'db.schema.table_a' → 'table_a'."""
    m = re.fullmatch(r'(?:\w+\.)+(\w+)', name.strip())
    return (m.group(1) if m else name.strip()).lower()


def clean_sql(raw_sql: str) -> str:
    """
    Pre-process raw SQL for reliable regex extraction:
    1. Remove block comments (/* ... */)
    2. Remove inline comments (-- ...)
    3. Collapse all whitespace to a single space
    """
    sql = RE_BLOCK_COMMENT.sub(' ', raw_sql)
    sql = RE_LINE_COMMENT.sub(' ', sql)
    sql = RE_WHITESPACE.sub(' ', sql)
    return sql.strip()


def extract_service_prefix(user_name: str) -> str | None:
    """Return the first matching SERVICE_PREFIX for a user_name, or None."""
    upper = user_name.upper()
    for prefix in SERVICE_PREFIXES:
        if upper.startswith(prefix):
            return prefix
    return None


def parse_joins(sql: str) -> list[tuple[str, str]]:
    """
    Extract directed (left_table, right_table) join pairs from cleaned SQL.

    Algorithm:
    1. Build an alias → physical_table map by scanning all FROM/JOIN refs.
    2. Walk JOIN clauses in source order.
    3. For each JOIN, the "left" table is the last physical table seen
       before that JOIN's position in the SQL string.
    4. Resolve all names through the alias map and strip schema prefixes.
    5. Skip pairs where either side is a SQL keyword or looks malformed.
    """

    # Step 1: Build alias map
    alias_map: dict[str, str] = {}
    positioned: list[tuple[int, str]] = []   # (char_position, phys_table_name)

    for m in RE_TABLE_REF.finditer(sql):
        raw_tbl   = m.group(1)
        raw_alias = m.group(2)
        phys      = strip_prefix(raw_tbl)

        if phys.upper() in SQL_KEYWORDS or not phys:
            continue

        alias_map[phys] = phys
        positioned.append((m.start(), phys))

        if raw_alias and raw_alias.upper() not in SQL_KEYWORDS:
            alias_map[raw_alias.strip().lower()] = phys

    def resolve(name: str) -> str:
        cleaned = strip_prefix(name)
        return alias_map.get(cleaned, cleaned)

    # Step 2: Walk JOIN clauses
    join_pairs: list[tuple[str, str]] = []

    for jm in RE_JOIN_CLAUSE.finditer(sql):
        join_start = jm.start()
        right_raw  = jm.group(1)
        right_phys = resolve(right_raw)

        if not right_phys or right_phys.upper() in SQL_KEYWORDS:
            continue

        # Find the latest table in `positioned` that appears before this JOIN
        # and is not the same physical table as right_phys
        left_phys = None
        for pos, tbl in reversed(positioned):
            if pos < join_start:
                candidate = resolve(tbl)
                if candidate != right_phys:
                    left_phys = candidate
                    break

        if not left_phys or left_phys.upper() in SQL_KEYWORDS:
            logger.debug("Could not resolve left table for JOIN at pos %d", join_start)
            continue

        join_pairs.append((left_phys, right_phys))

    return join_pairs


def compute_date_wid(log_date_str: str) -> str:
    """Convert a LogDate string (M/D/YYYY) to YYYYMMDD integer-as-string."""
    try:
        dt = pd.to_datetime(log_date_str, errors='raise', dayfirst=False)
        return dt.strftime('%Y%m%d')
    except Exception:
        return ''


def compute_metric_date(log_date_str: str) -> str:
    """Convert a LogDate string to YYYY-MM-DD format."""
    try:
        dt = pd.to_datetime(log_date_str, errors='raise', dayfirst=False)
        return dt.strftime('%Y-%m-%d')
    except Exception:
        return ''


def safe_duration_seconds(start: str, end: str) -> float | None:
    """Compute (end − start) in seconds; returns None on any failure."""
    try:
        s = pd.to_datetime(start, errors='raise')
        e = pd.to_datetime(end,   errors='raise')
        return (e - s).total_seconds()
    except Exception:
        return None


# ---------------------------------------------------------------------------
# INGESTION: RECORD BOUNDARY DETECTION
# ---------------------------------------------------------------------------

def load_records(filepath: str) -> tuple[list[str], int]:
    """
    Read the raw CSV file and split it into per-record text blocks.

    The file format is:
      - A new RECORD starts whenever a line begins with a compact username
        token (alphanumeric, no spaces) followed by a comma.
      - SQL text is multi-line and may contain blank lines — these are NOT
        record boundaries.
      - Standard CSV quoting is NOT reliably applied to the SQL field.

    Returns
    -------
    records_raw : list of raw text strings, one per record (header excluded)
    total       : count of data records found
    """
    logger.info("Reading input file: %s", filepath)
    try:
        with open(filepath, encoding='utf-8-sig') as fh:
            raw = fh.read()
    except FileNotFoundError:
        logger.critical("Input file not found: %s", filepath)
        sys.exit(1)
    except PermissionError:
        logger.critical("Permission denied reading: %s", filepath)
        sys.exit(1)

    lines = raw.split('\n')

    records_raw: list[str] = []
    current: list[str] = []

    for line in lines:
        if RE_RECORD_START.match(line) and current:
            records_raw.append('\n'.join(current))
            current = [line]
        else:
            current.append(line)

    if current:
        records_raw.append('\n'.join(current))

    # Remove header row (starts with "user_name")
    records_raw = [r for r in records_raw if not r.lstrip().startswith('user_name')]

    total = len(records_raw)
    logger.info("Detected %d data records.", total)
    return records_raw, total


# ---------------------------------------------------------------------------
# FIELD EXTRACTION FROM RAW RECORD TEXT
# ---------------------------------------------------------------------------

def parse_record_text(rec_text: str) -> tuple[dict | None, str | None]:
    """
    Extract the 7 fields from a raw record text block.

    Field layout (all on one logical "line" with SQL potentially spanning
    multiple physical lines):
        user_name , Db_nm , Tbl_nm , <SQL...> , LogDate , StartTime , LastResponseTime

    Strategy:
    1.  Collapse multiple consecutive blank lines to a single newline so the
        record is compact.
    2.  Split on the first 3 commas to extract user_name, Db_nm, Tbl_nm.
        These three fields are simple tokens that never contain commas.
    3.  In the remainder, find the first occurrence of a ",DATE" pattern
        (comma + date token) — this marks the boundary between SQL and the
        trailing timestamp fields.
    4.  Everything before the first ",DATE" is SqlTextInfo.
    5.  The tail after the first ",DATE" gives LogDate, StartTime,
        LastResponseTime (merging across any stray newlines).

    Returns (field_dict, None) on success, (None, error_reason) on failure.
    """
    # Collapse multi-blank-line gaps to single newline
    text = re.sub(r'\n{2,}', '\n', rec_text).strip()

    # --- Step 2: split first 3 fields ---
    parts = text.split(',', 3)
    if len(parts) < 4:
        return None, f"Less than 4 comma-separated segments found (got {len(parts)})"

    user_name = parts[0].strip()
    db_nm     = parts[1].strip()
    tbl_nm    = parts[2].strip()
    remainder = parts[3].strip()

    # Basic sanity: user_name should be a compact alphanumeric token
    if not re.fullmatch(r'[A-Za-z0-9]+', user_name):
        return None, f"user_name '{user_name}' does not look like a valid token"

    # --- Step 3: locate first ",DATE" boundary ---
    date_matches = list(RE_DATE_SEP.finditer(remainder))
    if not date_matches:
        return None, "No date/time fields found after SQL text"

    first_date_pos = date_matches[0].start()

    # --- Step 4: extract SQL text ---
    sql_text = remainder[:first_date_pos].strip()

    # Strip surrounding double-quotes that some rows use to wrap SQL
    if sql_text.startswith('"') and sql_text.endswith('"'):
        sql_text = sql_text[1:-1]
    elif sql_text.startswith('"'):
        sql_text = sql_text[1:]

    # --- Step 5: extract trailing timestamp fields ---
    tail = remainder[first_date_pos:]
    # Split the tail on commas; timestamps can contain spaces so we get them
    tail_parts = [p.strip() for p in tail.split(',') if p.strip()]

    if len(tail_parts) < 3:
        # LastResponseTime may be split across a newline (e.g. "3/30/2026\n2:35")
        # Try joining all tail_parts beyond index 1 as the final field
        if len(tail_parts) == 2:
            # Attempt to recover by checking if there's a stray time token
            tail_clean = re.sub(r'\s+', ' ', tail).strip()
            tail_parts2 = [p.strip() for p in tail_clean.split(',') if p.strip()]
            if len(tail_parts2) >= 3:
                tail_parts = tail_parts2
            else:
                return None, f"Only {len(tail_parts)} tail fields found (need 3): {tail_parts}"
        else:
            return None, f"Only {len(tail_parts)} tail fields found (need 3): {tail_parts}"

    log_date   = tail_parts[0]
    start_time = tail_parts[1]
    # LastResponseTime: join any remaining parts (handles newline-split case)
    last_resp  = re.sub(r'\s+', ' ', ' '.join(tail_parts[2:])).strip()

    return {
        'user_name':        user_name,
        'Db_nm':            db_nm,
        'Tbl_nm':           tbl_nm,
        'SqlTextInfo':      sql_text,
        'LogDate':          log_date,
        'StartTime':        start_time,
        'LastResponseTime': last_resp,
    }, None


# ---------------------------------------------------------------------------
# CORE PROCESSING
# ---------------------------------------------------------------------------

def process_records(
    records_raw: list[str],
) -> tuple[pd.DataFrame, int, int]:
    """
    Parse each raw record text into fields, extract join pairs, and
    assemble a flat DataFrame of raw join events.

    Each row in the output represents ONE join pair occurrence from ONE query.
    Aggregation to [left, right] grain happens in aggregate_join_pairs().

    Returns
    -------
    events_df           : raw join events DataFrame
    successfully_parsed : records that yielded ≥ 1 join pair
    skipped             : records skipped due to parse or extraction errors
    """
    raw_events  = []
    parsed_ok   = 0
    skipped     = 0

    for idx, rec_text in enumerate(records_raw):
        rec_num = idx + 2  # approximate CSV row number (1=header, 2=first data row)

        try:
            # ---- Field extraction ----------------------------------------
            fields, err = parse_record_text(rec_text)
            if err:
                logger.warning("SKIPPED record ~row %d – field parse error: %s", rec_num, err)
                skipped += 1
                continue

            user_name = fields['user_name']
            sql_raw   = fields['SqlTextInfo']
            log_date  = fields['LogDate']
            start_t   = fields['StartTime']
            end_t     = fields['LastResponseTime']

            # ---- SQL cleaning & join extraction --------------------------
            sql_clean = clean_sql(sql_raw)
            if not sql_clean:
                logger.debug("Record ~row %d – SQL is empty after cleaning; skipping.", rec_num)
                skipped += 1
                continue

            pairs = parse_joins(sql_clean)
            if not pairs:
                logger.debug("Record ~row %d – no join pairs found.", rec_num)
                skipped += 1
                continue

            # ---- Derive metrics ------------------------------------------
            date_wid    = compute_date_wid(log_date)
            metric_date = compute_metric_date(log_date)
            duration    = safe_duration_seconds(start_t, end_t)
            app_prefix  = extract_service_prefix(user_name)
            query_id    = idx   # unique identifier for this query within the run

            for left_tbl, right_tbl in pairs:
                raw_events.append({
                    'left_join_table_name':  left_tbl,
                    'right_join_table_name': right_tbl,
                    'date_wid':              date_wid,
                    'metric_date':           metric_date,
                    'user_name':             user_name,
                    'app_prefix':            app_prefix,
                    'query_id':              query_id,
                    'duration_sec':          duration,
                })

            parsed_ok += 1

        except Exception as exc:
            logger.warning(
                "SKIPPED record ~row %d – unexpected error: %s",
                rec_num, exc, exc_info=True,
            )
            skipped += 1

    cols = [
        'left_join_table_name', 'right_join_table_name',
        'date_wid', 'metric_date', 'user_name', 'app_prefix',
        'query_id', 'duration_sec',
    ]
    events_df = pd.DataFrame(raw_events) if raw_events else pd.DataFrame(columns=cols)

    logger.info(
        "Processing complete – Parsed OK: %d | Skipped: %d | Raw join events: %d",
        parsed_ok, skipped, len(events_df),
    )
    return events_df, parsed_ok, skipped


# ---------------------------------------------------------------------------
# AGGREGATION
# ---------------------------------------------------------------------------

def aggregate_join_pairs(events_df: pd.DataFrame) -> pd.DataFrame:
    """
    Aggregate raw join events to the [left_table, right_table, date] grain.

    Output columns (exact required order):
        Row_id, date_wid, Metric date, left_join_table_name,
        right_join_table_name, join_count, unique_users, unique_app,
        query_count, avg_runtime
    """
    if events_df.empty:
        logger.warning("No join events to aggregate – output will be empty.")
        return pd.DataFrame()

    group_keys = [
        'left_join_table_name', 'right_join_table_name',
        'date_wid', 'metric_date',
    ]
    agg_records = []

    for group_vals, grp in events_df.groupby(group_keys, sort=True):
        left_tbl, right_tbl, date_wid, metric_date = group_vals

        join_count   = len(grp)
        unique_users = grp['user_name'].nunique()
        unique_app   = grp['app_prefix'].dropna().nunique()
        query_count  = grp['query_id'].nunique()

        valid_dur    = grp['duration_sec'].dropna()
        avg_runtime  = round(float(valid_dur.mean()), 4) if not valid_dur.empty else None

        agg_records.append({
            'date_wid':               date_wid,
            'Metric date':            metric_date,
            'left_join_table_name':   left_tbl,
            'right_join_table_name':  right_tbl,
            'join_count':             join_count,
            'unique_users':           unique_users,
            'unique_app':             unique_app,
            'query_count':            query_count,
            'avg_runtime':            avg_runtime,
        })

    agg_df = pd.DataFrame(agg_records)
    agg_df.insert(0, 'Row_id', range(1, len(agg_df) + 1))

    # Enforce exact column order per specification
    final_cols = [
        'Row_id', 'date_wid', 'Metric date',
        'left_join_table_name', 'right_join_table_name',
        'join_count', 'unique_users', 'unique_app',
        'query_count', 'avg_runtime',
    ]
    agg_df = agg_df[final_cols]

    logger.info("Aggregation complete – %d unique join pairs in output.", len(agg_df))
    return agg_df


# ---------------------------------------------------------------------------
# EXCEL OUTPUT
# ---------------------------------------------------------------------------

def write_excel(agg_df: pd.DataFrame, output_path: str) -> None:
    """Write the aggregated DataFrame to a professionally formatted Excel file."""

    wb = Workbook()
    ws = wb.active
    ws.title = "Join Analysis"

    # Styles
    header_font  = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    header_fill  = PatternFill("solid", start_color="2F5496")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    data_font    = Font(name="Arial", size=10)
    alt_fill     = PatternFill("solid", start_color="EEF2FA")
    center_align = Alignment(horizontal="center", vertical="center")
    left_align   = Alignment(horizontal="left",   vertical="center")

    thin  = Side(style="thin", color="B8C4D8")
    bdr   = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers = list(agg_df.columns)
    ws.append(headers)
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = header_align
        cell.border    = bdr
    ws.row_dimensions[1].height = 28

    for row_idx, df_row in enumerate(agg_df.itertuples(index=False), start=2):
        ws.append(list(df_row))
        fill = alt_fill if row_idx % 2 == 0 else None

        for col_idx in range(1, len(headers) + 1):
            cell      = ws.cell(row=row_idx, column=col_idx)
            cell.font = data_font
            cell.border = bdr
            if fill:
                cell.fill = fill

            col_name = headers[col_idx - 1]
            if col_name in ('left_join_table_name', 'right_join_table_name', 'Metric date'):
                cell.alignment = left_align
            else:
                cell.alignment = center_align

            if col_name == 'avg_runtime':
                cell.number_format = '0.0000'
            elif col_name in ('join_count', 'unique_users', 'unique_app', 'query_count', 'Row_id'):
                cell.number_format = '#,##0'

        ws.row_dimensions[row_idx].height = 18

    col_widths = {
        'Row_id': 8, 'date_wid': 12, 'Metric date': 14,
        'left_join_table_name': 32, 'right_join_table_name': 32,
        'join_count': 12, 'unique_users': 14,
        'unique_app': 12, 'query_count': 14, 'avg_runtime': 14,
    }
    for col_idx, col_name in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = col_widths.get(col_name, 15)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    wb.save(output_path)
    logger.info("Excel output saved: %s", output_path)


# ---------------------------------------------------------------------------
# SUMMARY
# ---------------------------------------------------------------------------

def print_summary(total: int, parsed: int, skipped: int) -> None:
    line = "=" * 55
    print(f"\n{line}")
    print("  PROCESSING SUMMARY")
    print(line)
    print(f"  Total Queries Processed       : {total:>8,}")
    print(f"  Total Queries Successfully    : {parsed:>8,}")
    print(f"  Total Queries Skipped         : {skipped:>8,}")
    print(line)


# ---------------------------------------------------------------------------
# MAIN
# ---------------------------------------------------------------------------

def main() -> None:
    parser = argparse.ArgumentParser(
        description=(
            "Parse SQL join relationships from a multi-line query log CSV "
            "and export a join-pair analysis to Excel."
        )
    )
    parser.add_argument('--input',  '-i', required=True,
                        help="Path to the input CSV file.")
    parser.add_argument('--output', '-o', default="join_analysis_output.xlsx",
                        help="Path for the output Excel file (default: join_analysis_output.xlsx).")
    args = parser.parse_args()

    logger.info("=== SQL Join Parser – START ===")
    logger.info("Input  : %s", args.input)
    logger.info("Output : %s", args.output)

    records_raw, total        = load_records(args.input)
    events_df, parsed, skipped = process_records(records_raw)
    agg_df                    = aggregate_join_pairs(events_df)

    if not agg_df.empty:
        write_excel(agg_df, args.output)
    else:
        logger.warning("No join pairs found – output file will NOT be created.")

    print_summary(total, parsed, skipped)
    logger.info("=== SQL Join Parser – DONE ===")


if __name__ == "__main__":
    main()