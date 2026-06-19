"""
sql_join_lineage.py
===================
Production-grade Teradata SQL log parser.

Reads an Excel/CSV file containing raw SQL query logs and outputs a detailed,
row-by-row breakdown of table-join relationships into a new Excel file.

Input columns (case-insensitive):
  user_name, Db_nm, Tbl_nm, SqlTextInfo, LogDate, StartTime, LastResponseTime

Output columns (exact order):
  Row_id, date_wid, Metric date, left_join_table_name, right_join_table_name,
  join_count, unique_users, unique_app, query_count, avg_runtime

Usage:
  python sql_join_lineage.py input.xlsx
  python sql_join_lineage.py input.xlsx -o output.xlsx --sheet "Sheet1" --debug
"""

from __future__ import annotations

import argparse
import logging
import re
import warnings
from pathlib import Path
from typing import Optional

import pandas as pd
import sqlglot
from sqlglot import exp

warnings.filterwarnings("ignore")

# ---------------------------------------------------------------------------
# Logging
# ---------------------------------------------------------------------------
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# App-prefix classifier
# ---------------------------------------------------------------------------

APP_PREFIXES = ("SVP", "SVR", "SVC", "SVT", "ETL", "DBA", "OP")

def classify_app(user_name: str) -> str:
    """Return the matching app prefix or 'OTHER'."""
    uname = str(user_name).strip().upper()
    for prefix in APP_PREFIXES:
        if uname.startswith(prefix):
            return prefix
    return "OTHER"


# ---------------------------------------------------------------------------
# Date/time helpers
# ---------------------------------------------------------------------------

def make_date_wid(date_val) -> str:
    """YYYYMMDD integer-like string for date_wid."""
    try:
        return pd.to_datetime(date_val).strftime("%Y%m%d")
    except Exception:
        return "00000000"


def make_metric_date(date_val) -> str:
    """Clean calendar date string YYYY-MM-DD for Metric date."""
    try:
        return pd.to_datetime(date_val).strftime("%Y-%m-%d")
    except Exception:
        return ""


def compute_runtime_seconds(start, end) -> Optional[float]:
    """Return (end - start) in seconds, or None if either value is unparseable."""
    try:
        s = pd.to_datetime(start)
        e = pd.to_datetime(end)
        delta = (e - s).total_seconds()
        return max(delta, 0.0)          # guard against negative clock skew
    except Exception:
        return None


# ---------------------------------------------------------------------------
# Name normalisation
# ---------------------------------------------------------------------------

def normalize_name(name: Optional[str]) -> str:
    if not name:
        return ""
    return str(name).strip().strip('"').strip("'").lower()


def strip_prefix(qualified_name: str) -> str:
    """
    Extract the bare table name from any db.schema.table or db.table form.
    'mydb.myschema.orders' → 'orders'
    'mydb.orders'          → 'orders'
    'orders'               → 'orders'
    """
    if not qualified_name:
        return ""
    parts = qualified_name.strip().split(".")
    return parts[-1].strip()


def qualify(db: Optional[str], tbl: Optional[str]) -> str:
    """Return db.table or just table (used internally before stripping prefix)."""
    db  = normalize_name(db)
    tbl = normalize_name(tbl)
    if db and tbl:
        return f"{db}.{tbl}"
    return tbl or db or ""


# ---------------------------------------------------------------------------
# Teradata SQL pre-cleaner
# ---------------------------------------------------------------------------

# Each tuple: (pattern, replacement[, flags])
TERADATA_CLEANUP: list[tuple] = [
    # Teradata-specific clauses
    (r"\bFORMAT\s+'[^']*'",                                                 ""),
    (r"\bTITLE\s+'[^']*'",                                                  ""),
    (r"\bNAMED\s+\w+",                                                      ""),
    (r"\bCASEWORD\s+\w+",                                                   ""),
    (r"\bAT\s+ISOLATION\s+LEVEL\s+\w+",                                     ""),
    (r"\bLOCKING\s+(?:ROW|TABLE|DATABASE|VIEW)\s+(?:FOR\s+)?(?:ACCESS|WRITE|READ|EXCLUSIVE)", ""),
    (r"\bLOCKING\s+\S+\s+FOR\s+(?:ACCESS|WRITE|READ|EXCLUSIVE)",           ""),
    (r"\b(?:COMPRESS|NO\s+COMPRESS)\b[^,;)]*",                              ""),
    (r"\bNORMALIZE(?:\s+ON\s+MEETS\s+OR\s+OVERLAPS)?",                     ""),
    (r"\bEXPAND\s+ON\s+\w+",                                                ""),
    # Comments (single-line and block)
    (r"--[^\n]*",                                                            ""),
    (r"/\*.*?\*/",                                                           " ", re.DOTALL),
    # BTEQ artefacts and trailing semi-colon
    (r"\bBTEQ\b.*",                                                         ""),
    (r";\s*$",                                                               ""),
    # Collapse whitespace
    (r"\s+",                                                                 " "),
]


def clean_sql(sql: str) -> str:
    """Strip Teradata-specific syntax noise and comments from a SQL string."""
    s = str(sql).strip()
    for item in TERADATA_CLEANUP:
        flags = item[2] if len(item) == 3 else 0
        s = re.sub(item[0], item[1], s, flags=flags | re.IGNORECASE)
    return s.strip()


# ---------------------------------------------------------------------------
# Regex-based fallback join extractor
# ---------------------------------------------------------------------------

# Matches FROM or JOIN + optional db prefix + table name + optional alias
_JOIN_TABLE_RE = re.compile(
    r"""
    (?:FROM|JOIN)\s+
    (?:([\w$]+)\s*\.\s*)?([\w$]+)        # optional db + table
    (?:\s+AS\s+([\w$]+)|\s+([\w$]+))?    # optional alias
    """,
    re.IGNORECASE | re.VERBOSE,
)

# Matches an explicit JOIN clause (with type) and its ON condition
_JOIN_CLAUSE_RE = re.compile(
    r"""
    (?:(?:LEFT|RIGHT|FULL|CROSS|INNER|OUTER)\s+(?:OUTER\s+)?)?
    JOIN\s+
    (?:([\w$]+)\s*\.\s*)?([\w$]+)        # optional db + table
    (?:\s+AS\s+([\w$]+)|\s+([\w$]+))?    # optional alias
    \s+ON\s+(.+?)                         # ON clause content
    (?=
        (?:LEFT|RIGHT|FULL|CROSS|INNER|OUTER)?\s*JOIN
        |\s+WHERE|\s+GROUP|\s+ORDER|\s+HAVING
        |$
    )
    """,
    re.IGNORECASE | re.VERBOSE | re.DOTALL,
)

_ON_KEYS_RE = re.compile(r"([\w$.]+)\s*=\s*([\w$.]+)", re.IGNORECASE)


def _resolve_alias(alias_map: dict, ref: str) -> str:
    ref = normalize_name(ref)
    return alias_map.get(ref, ref)


def regex_extract_joins(sql: str, fallback_db: str = "") -> list[dict]:
    """
    Pure-regex fallback.  Returns a list of dicts:
      { left_table: str, right_table: str, join_keys: list[dict] }
    """
    alias_map: dict[str, str] = {}
    tables_in_order: list[str] = []

    # First pass: build alias map and ordered table list
    for m in _JOIN_TABLE_RE.finditer(sql):
        db    = m.group(1) or fallback_db
        tbl   = m.group(2)
        alias = m.group(3) or m.group(4) or tbl
        q     = qualify(db, tbl)
        alias_map[normalize_name(alias)] = q
        alias_map[normalize_name(tbl)]   = q
        if q not in tables_in_order:
            tables_in_order.append(q)

    joins: list[dict] = []

    # Second pass: extract per-JOIN ON-condition details
    for m in _JOIN_CLAUSE_RE.finditer(sql):
        db          = m.group(1) or fallback_db
        tbl         = m.group(2)
        alias       = m.group(3) or m.group(4) or tbl
        on_clause   = m.group(5)
        right_table = qualify(db, tbl)
        alias_map[normalize_name(alias)] = right_table

        keys: list[dict] = []
        left_table: Optional[str] = None

        for lk, rk in _ON_KEYS_RE.findall(on_clause):
            l_parts = lk.split(".")
            r_parts = rk.split(".")
            l_tbl = _resolve_alias(alias_map, l_parts[0]) if len(l_parts) > 1 else None
            r_tbl = _resolve_alias(alias_map, r_parts[0]) if len(r_parts) > 1 else None
            l_col = l_parts[-1]
            r_col = r_parts[-1]

            if r_tbl and r_tbl == right_table:
                left_table = l_tbl
                keys.append({"left_key": l_col, "right_key": r_col})
            elif l_tbl and l_tbl == right_table:
                left_table = r_tbl
                keys.append({"left_key": r_col, "right_key": l_col})
            else:
                keys.append({"left_key": l_col, "right_key": r_col})

        # If ON clause didn't resolve the left side, infer from table order
        if not left_table and len(tables_in_order) >= 2:
            try:
                idx = tables_in_order.index(right_table)
                left_table = tables_in_order[idx - 1] if idx > 0 else tables_in_order[0]
            except ValueError:
                left_table = tables_in_order[0] if tables_in_order else ""

        joins.append({
            "left_table":  left_table or "",
            "right_table": right_table,
            "join_keys":   keys,
        })

    return joins


# ---------------------------------------------------------------------------
# AST-based join extractor (sqlglot)
# ---------------------------------------------------------------------------

def ast_extract_joins(sql: str, fallback_db: str = "") -> list[dict]:
    """
    Parse SQL with sqlglot (Teradata dialect) and walk the AST to extract
    sequential join pairs.  Returns same shape as regex_extract_joins.
    """
    try:
        statements = sqlglot.parse(
            sql,
            dialect="teradata",
            error_level=sqlglot.ErrorLevel.IGNORE,
        )
    except Exception as exc:
        log.debug("sqlglot.parse failed: %s", exc)
        return []

    if not statements:
        return []

    joins_out: list[dict] = []

    for stmt in statements:
        if stmt is None:
            continue

        # ── Build alias map for this statement ──────────────────────────────
        alias_map: dict[str, str] = {}
        for tbl_expr in stmt.find_all(exp.Table):
            db_node    = tbl_expr.args.get("db")
            name_node  = tbl_expr.args.get("this")
            alias_node = tbl_expr.args.get("alias")
            db_s    = db_node.name  if db_node    else fallback_db
            tbl_s   = name_node.name if name_node else ""
            alias_s = alias_node.name if alias_node else tbl_s
            q = qualify(db_s, tbl_s)
            alias_map[normalize_name(alias_s)] = q
            alias_map[normalize_name(tbl_s)]   = q

        def resolve_node(node) -> str:
            if isinstance(node, exp.Table):
                db_n   = node.args.get("db")
                name_n = node.args.get("this")
                return qualify(
                    db_n.name  if db_n   else fallback_db,
                    name_n.name if name_n else "",
                )
            if isinstance(node, exp.Column):
                tbl = node.args.get("table")
                if tbl:
                    return alias_map.get(normalize_name(tbl.name), tbl.name)
            return ""

        # ── Walk every SELECT block ──────────────────────────────────────────
        for select in stmt.find_all(exp.Select):
            from_clause = select.args.get("from")
            if not from_clause:
                continue
            from_tbl = from_clause.find(exp.Table)
            if from_tbl is None:
                continue
            left_anchor = resolve_node(from_tbl)

            for join_node in select.args.get("joins", []):
                right_expr  = join_node.args.get("this")
                right_table = resolve_node(right_expr) if right_expr else ""

                on_node    = join_node.args.get("on")
                join_keys: list[dict] = []
                if on_node:
                    for eq in on_node.find_all(exp.EQ):
                        lc = eq.args.get("this")
                        rc = eq.args.get("expression")
                        if lc and rc:
                            join_keys.append({
                                "left_key":  lc.name if hasattr(lc, "name") else str(lc),
                                "right_key": rc.name if hasattr(rc, "name") else str(rc),
                            })

                joins_out.append({
                    "left_table":  left_anchor,
                    "right_table": right_table,
                    "join_keys":   join_keys,
                })
                # Roll anchor forward for chained joins: A→B, B→C, C→D …
                left_anchor = right_table or left_anchor

    return joins_out


# ---------------------------------------------------------------------------
# Master join extractor (AST → regex fallback → dedup)
# ---------------------------------------------------------------------------

def extract_joins(sql: str, fallback_db: str = "") -> list[dict]:
    """
    Try AST-based extraction first; fall back to regex for unparseable SQL.
    Returns deduplicated list of { left_table, right_table, join_keys }.
    Table names in left_table / right_table are **bare** names (prefix stripped).
    """
    cleaned = clean_sql(sql)
    joins: list[dict] = []

    try:
        joins = ast_extract_joins(cleaned, fallback_db)
    except Exception as exc:
        log.debug("AST extraction raised (%s), switching to regex.", exc)

    if not joins:
        log.debug("AST found no joins; trying regex fallback.")
        try:
            joins = regex_extract_joins(cleaned, fallback_db)
        except Exception as exc:
            log.debug("Regex extraction also failed: %s", exc)

    # Strip db/schema prefix → bare table name, deduplicate, drop incomplete pairs
    seen: set[tuple] = set()
    unique: list[dict] = []
    for j in joins:
        left  = strip_prefix(j.get("left_table",  ""))
        right = strip_prefix(j.get("right_table", ""))
        if not left or not right:
            continue
        key = (left, right)
        if key not in seen:
            seen.add(key)
            unique.append({
                "left_table":  left,
                "right_table": right,
                "join_keys":   j.get("join_keys", []),
            })
    return unique


# ---------------------------------------------------------------------------
# Core transformation
# ---------------------------------------------------------------------------

def transform_logs(df: pd.DataFrame) -> pd.DataFrame:
    """
    Accept raw log DataFrame; return the 10-column aggregated output DataFrame.

    Output column order (strict):
      Row_id, date_wid, Metric date,
      left_join_table_name, right_join_table_name,
      join_count, unique_users, unique_app, query_count, avg_runtime
    """
    # ── Flexible column resolution (case-insensitive) ───────────────────────
    col_lower = {c.lower(): c for c in df.columns}

    REQUIRED = ["sqltextinfo"]
    OPTIONAL_DEFAULTS: dict[str, str] = {
        "user_name":       "",
        "db_nm":           "",
        "tbl_nm":          "",
        "logdate":         "",
        "starttime":       "",
        "lastresponsetime":"",
    }

    for req in REQUIRED:
        if req not in col_lower:
            raise ValueError(f"Required column missing: '{req}' (case-insensitive).  "
                             f"Found: {list(df.columns)}")

    def get_series(name: str, default="") -> pd.Series:
        actual = col_lower.get(name.lower())
        if actual:
            return df[actual].fillna(default).astype(str)
        log.warning("Column '%s' not found – using default '%s'.", name, default)
        return pd.Series([default] * len(df), dtype=str)

    col_sql       = get_series("sqltextinfo")
    col_user      = get_series("user_name")
    col_db        = get_series("db_nm")
    col_tbl       = get_series("tbl_nm")
    col_logdate   = df[col_lower["logdate"]]   if "logdate"            in col_lower else pd.Series([""] * len(df))
    col_starttime = df[col_lower["starttime"]] if "starttime"          in col_lower else pd.Series([""] * len(df))
    col_endtime   = df[col_lower["lastresponsetime"]] if "lastresponsetime" in col_lower else pd.Series([""] * len(df))

    # ── Pass 1: per-query join-pair explosion ────────────────────────────────
    exploded: list[dict] = []
    total = len(df)

    for i in range(total):
        sql  = str(col_sql.iloc[i]).strip()
        usr  = str(col_user.iloc[i]).strip()
        db   = str(col_db.iloc[i]).strip()
        tbl  = str(col_tbl.iloc[i]).strip()
        date = col_logdate.iloc[i]
        t0   = col_starttime.iloc[i]
        t1   = col_endtime.iloc[i]

        if not sql or sql.lower() in ("nan", "none", ""):
            log.debug("Row %d: empty SQL, skipped.", i)
            continue

        runtime = compute_runtime_seconds(t0, t1)
        app_tag = classify_app(usr)

        try:
            joins = extract_joins(sql, fallback_db=db)
        except Exception as exc:
            log.warning("Row %d: join extraction failed (%s), skipping.", i, exc)
            continue

        if not joins:
            # Queries with no JOIN: record with driving table only (right = "")
            driving = strip_prefix(qualify(db, tbl))
            if driving:
                exploded.append({
                    "left_table":  driving,
                    "right_table": "",
                    "user":        usr,
                    "app":         app_tag,
                    "date":        date,
                    "runtime":     runtime,
                })
            continue

        for j in joins:
            exploded.append({
                "left_table":  j["left_table"],
                "right_table": j["right_table"],
                "user":        usr,
                "app":         app_tag,
                "date":        date,
                "runtime":     runtime,
            })

        if (i + 1) % 500 == 0:
            log.info("Processed %d / %d rows …", i + 1, total)

    log.info("Pass 1 complete: %d exploded join-pair records.", len(exploded))

    if not exploded:
        log.warning("No join pairs found.  Output will be empty.")
        return pd.DataFrame()

    raw = pd.DataFrame(exploded)

    # ── Derive date keys ─────────────────────────────────────────────────────
    raw["date_wid"]     = raw["date"].apply(make_date_wid)
    raw["Metric date"]  = raw["date"].apply(make_metric_date)

    # ── Pass 2: aggregation ──────────────────────────────────────────────────
    GROUP_KEYS = ["date_wid", "Metric date", "left_table", "right_table"]

    agg_rows: list[dict] = []

    for group_key, grp in raw.groupby(GROUP_KEYS, sort=True):
        dw, md, left, right = group_key

        join_count   = len(grp)
        unique_users = grp["user"].nunique()
        unique_app   = grp["app"].nunique()
        query_count  = len(grp)   # one record = one query hit; same as join_count here

        runtimes = grp["runtime"].dropna()
        avg_rt   = round(runtimes.mean(), 3) if not runtimes.empty else None

        agg_rows.append({
            "date_wid":               dw,
            "Metric date":            md,
            "left_join_table_name":   left,
            "right_join_table_name":  right if right else "",
            "join_count":             join_count,
            "unique_users":           unique_users,
            "unique_app":             unique_app,
            "query_count":            query_count,
            "avg_runtime":            avg_rt,
        })

    out = pd.DataFrame(agg_rows)

    # ── Post-aggregation: sort then add sequential Row_id ───────────────────
    out.sort_values(
        by=["date_wid", "left_join_table_name", "right_join_table_name"],
        inplace=True,
    )
    out.reset_index(drop=True, inplace=True)
    out.insert(0, "Row_id", range(1, len(out) + 1))

    # ── Enforce exact column order ───────────────────────────────────────────
    FINAL_COLS = [
        "Row_id",
        "date_wid",
        "Metric date",
        "left_join_table_name",
        "right_join_table_name",
        "join_count",
        "unique_users",
        "unique_app",
        "query_count",
        "avg_runtime",
    ]
    out = out[FINAL_COLS]

    log.info("Pass 2 complete: %d aggregated rows.", len(out))
    return out


# ---------------------------------------------------------------------------
# I/O helpers
# ---------------------------------------------------------------------------

def read_input(path: str, sheet=0) -> pd.DataFrame:
    p   = Path(path)
    ext = p.suffix.lower()
    log.info("Reading input: %s", path)
    if ext in (".xlsx", ".xls", ".xlsm"):
        df = pd.read_excel(path, sheet_name=sheet, dtype=str)
    elif ext == ".csv":
        df = pd.read_csv(path, dtype=str)
    else:
        raise ValueError(f"Unsupported file type: {ext}  (expected .xlsx/.xls/.xlsm/.csv)")
    log.info("Loaded %d rows | columns: %s", len(df), list(df.columns))
    return df


def write_output(df: pd.DataFrame, out_path: str) -> None:
    p   = Path(out_path)
    ext = p.suffix.lower()
    p.parent.mkdir(parents=True, exist_ok=True)

    if ext in (".xlsx", ".xls", ".xlsm"):
        with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name="join_lineage")

            # ── Light auto-formatting ───────────────────────────────────────
            ws = writer.sheets["join_lineage"]
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.utils import get_column_letter

            HEADER_FILL  = PatternFill("solid", fgColor="1F4E79")
            HEADER_FONT  = Font(bold=True, color="FFFFFF", size=11)
            ALIGN_CENTER = Alignment(horizontal="center", vertical="center")

            for col_idx, cell in enumerate(ws[1], start=1):
                cell.fill      = HEADER_FILL
                cell.font      = HEADER_FONT
                cell.alignment = ALIGN_CENTER
                # Auto-width (capped at 50)
                col_width = max(
                    len(str(cell.value or "")),
                    *(len(str(ws.cell(row=r, column=col_idx).value or ""))
                      for r in range(2, min(ws.max_row + 1, 100))),
                )
                ws.column_dimensions[get_column_letter(col_idx)].width = min(col_width + 4, 50)

            ws.freeze_panes = "A2"
    else:
        df.to_csv(out_path, index=False)

    log.info("Output written → %s  (%d rows)", out_path, len(df))


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def build_arg_parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(
        description=(
            "Parse Teradata SQL query logs and produce a detailed "
            "row-by-row join-lineage Excel report."
        ),
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  python sql_join_lineage.py logs.xlsx
  python sql_join_lineage.py logs.xlsx -o reports/join_report.xlsx
  python sql_join_lineage.py logs.csv --sheet 0 --debug
        """,
    )
    p.add_argument("input",  help="Path to input Excel (.xlsx/.xls) or CSV file")
    p.add_argument(
        "-o", "--output",
        default=None,
        help="Output file path  (default: <input_stem>_join_lineage.xlsx)",
    )
    p.add_argument(
        "--sheet",
        default=0,
        help="Sheet name or 0-based index for Excel input  (default: 0)",
    )
    p.add_argument(
        "--debug",
        action="store_true",
        help="Enable DEBUG-level logging (verbose SQL parse traces)",
    )
    return p


def main() -> None:
    args = build_arg_parser().parse_args()

    if args.debug:
        logging.getLogger().setLevel(logging.DEBUG)

    # Resolve output path
    if args.output is None:
        p = Path(args.input)
        args.output = str(p.parent / f"{p.stem}_join_lineage.xlsx")

    # Sheet coercion: try int first
    sheet = args.sheet
    try:
        sheet = int(sheet)
    except (ValueError, TypeError):
        pass   # keep as string (sheet name)

    df_in  = read_input(args.input, sheet=sheet)
    df_out = transform_logs(df_in)

    if df_out.empty:
        log.warning("No data produced — check that SqlTextInfo contains JOIN queries.")
        return

    write_output(df_out, args.output)

    # ── Summary ─────────────────────────────────────────────────────────────
    print("\n" + "=" * 60)
    print(f"✅  Done!  {len(df_out):,} lineage rows written to:")
    print(f"   {args.output}")
    print("=" * 60)
    print("\nColumn order:", list(df_out.columns))
    print(f"\nTop 10 rows (of {len(df_out):,}):")
    print(df_out.head(10).to_string(index=False))
    print("=" * 60)


if __name__ == "__main__":
    main()
