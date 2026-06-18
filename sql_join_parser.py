"""
sql_join_parser.py
==================
Production-grade pipeline that parses database query logs from an Excel sheet
and outputs a granular, row-by-row breakdown of table join relationships.

Author  : Senior Data Engineer
Version : 1.0.0
Requires: pandas, sqlglot, openpyxl, xlsxwriter
"""

from __future__ import annotations

import argparse
import logging
import re
import sys
from pathlib import Path
from typing import Optional

import pandas as pd
import sqlglot
from sqlglot import exp
from openpyxl import load_workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    GradientFill,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Logging
# ---------------------------------------------------------------------------
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
)
logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------
APP_PREFIXES = ("SVP", "SVR", "SVC", "SVT", "ETL", "DBA", "OP")

INPUT_REQUIRED_COLS = {
    "user_name",
    "Db_nm",
    "Tbl_nm",
    "SqlTextInfo",
    "LogDate",
    "StartTime",
    "LastResponseTime",
}

OUTPUT_COLUMNS = [
    "Row_id",
    "date_wid",
    "Metric date",
    "left_join_table_name",
    "right_join_table_name",
    "join_count",
    "unique_users",
    "unique_app",
    "query_count",
    "avg_runtime",
]

# ---------------------------------------------------------------------------
# SQL Parsing helpers
# ---------------------------------------------------------------------------

def _strip_hints_and_comments(sql: str) -> str:
    """Remove inline hints (/*+ ... */) and standard block/line comments."""
    # Remove optimizer hints  /*+ ... */
    sql = re.sub(r"/\*\+.*?\*/", " ", sql, flags=re.DOTALL)
    # Remove block comments /* ... */
    sql = re.sub(r"/\*.*?\*/", " ", sql, flags=re.DOTALL)
    # Remove line comments -- ...
    sql = re.sub(r"--[^\n]*", " ", sql)
    return sql.strip()


def _is_inside_subquery(node: exp.Expression) -> bool:
    """Return True when *node* is a descendant of a Subquery expression."""
    parent = node.parent
    while parent is not None:
        if isinstance(parent, exp.Subquery):
            return True
        parent = parent.parent
    return False


def extract_join_pairs(
    sql: str,
    driving_table: Optional[str] = None,
) -> list[tuple[str, str]]:
    """
    Parse *sql* with sqlglot and return an ordered list of (left, right) join
    pairs extracted from the **top-level** query scope only.

    Strategy
    --------
    1. Strip comments/hints.
    2. Parse with sqlglot (dialect-agnostic, WARN on errors).
    3. Locate the top-level FROM table → first left-hand table.
    4. Walk the AST for Join nodes **not inside a Subquery**.
    5. Build sequential pairs: (from_tbl → join1), (join1 → join2), …
    6. Skip any Join whose right-hand side is a Subquery (no physical name).
    7. Resolve aliases back to real table names via the alias map built in step 3-4.

    Parameters
    ----------
    sql            : Raw SQL text from the log.
    driving_table  : Optional fallback table name from ``Tbl_nm`` column.

    Returns
    -------
    List of (left_table, right_table) string tuples (clean names only).
    """
    clean_sql = _strip_hints_and_comments(sql)
    if not clean_sql:
        return []

    try:
        tree = sqlglot.parse_one(clean_sql, error_level=sqlglot.ErrorLevel.WARN)
    except Exception as exc:  # noqa: BLE001
        raise ValueError(f"sqlglot parse failed: {exc}") from exc

    # ----- Build alias → real_name map from all tables in AST ----------------
    alias_map: dict[str, str] = {}
    for node in tree.walk():
        if isinstance(node, exp.Table) and node.name:
            real = node.name.strip()
            alias = node.alias.strip() if node.alias else ""
            if alias and alias.lower() != real.lower():
                alias_map[alias.lower()] = real

    def resolve(name: str) -> str:
        return alias_map.get(name.lower(), name)

    # ----- Collect top-level FROM and JOIN nodes (skip inside subqueries) ----
    from_table: Optional[str] = None
    join_table_sequence: list[str] = []  # ordered right-hand tables

    for node in tree.walk():
        if _is_inside_subquery(node):
            continue

        if isinstance(node, exp.From):
            tbl = node.find(exp.Table)
            if tbl and tbl.name:
                from_table = resolve(tbl.name.strip())

        elif isinstance(node, exp.Join):
            right = node.args.get("this")
            if isinstance(right, exp.Table) and right.name:
                join_table_sequence.append(resolve(right.name.strip()))
            # Subquery-based joins → no physical table; skip silently

    # Fallback: if FROM not found, try Tbl_nm driving table
    if from_table is None and driving_table:
        from_table = driving_table.strip().split(".")[-1]

    if from_table is None or not join_table_sequence:
        return []

    # ----- Build sequential pairs --------------------------------------------
    pairs: list[tuple[str, str]] = []
    left = from_table
    for right in join_table_sequence:
        if left and right and left.lower() != right.lower():
            pairs.append((left, right))
        left = right  # chain: next left is current right

    return pairs


# ---------------------------------------------------------------------------
# Application prefix extractor
# ---------------------------------------------------------------------------

def _extract_app_prefix(user_name: str) -> Optional[str]:
    """Return the matching APP_PREFIX for a user_name, or None."""
    if not isinstance(user_name, str):
        return None
    upper = user_name.upper()
    for prefix in APP_PREFIXES:
        if upper.startswith(prefix):
            return prefix
    return None


# ---------------------------------------------------------------------------
# Core pipeline
# ---------------------------------------------------------------------------

def load_and_validate_input(filepath: str) -> pd.DataFrame:
    """Read the Excel input file and validate required columns."""
    logger.info("Loading input file: %s", filepath)
    df = pd.read_excel(filepath, dtype={"user_name": str, "Db_nm": str, "Tbl_nm": str})

    missing = INPUT_REQUIRED_COLS - set(df.columns)
    if missing:
        raise ValueError(f"Input file is missing required columns: {missing}")

    logger.info("Loaded %d rows from input.", len(df))
    return df


def parse_dates(df: pd.DataFrame) -> pd.DataFrame:
    """Coerce date/time columns to proper types."""
    df["LogDate"] = pd.to_datetime(df["LogDate"], errors="coerce")
    df["StartTime"] = pd.to_datetime(df["StartTime"], errors="coerce")
    df["LastResponseTime"] = pd.to_datetime(df["LastResponseTime"], errors="coerce")

    null_dates = df["LogDate"].isna().sum()
    if null_dates:
        logger.warning("%d rows have unparseable LogDate; they will be skipped.", null_dates)
    df = df.dropna(subset=["LogDate"]).copy()

    df["date_wid"] = df["LogDate"].dt.strftime("%Y%m%d").astype(str)
    df["Metric date"] = df["LogDate"].dt.normalize()
    df["runtime_seconds"] = (
        df["LastResponseTime"] - df["StartTime"]
    ).dt.total_seconds().clip(lower=0)
    df["app_prefix"] = df["user_name"].apply(_extract_app_prefix)

    return df


def explode_join_pairs(df: pd.DataFrame) -> pd.DataFrame:
    """
    For every row in *df*, extract join pairs from SqlTextInfo and
    produce one expanded row per (query_row × join_pair).
    """
    records: list[dict] = []
    skipped = 0

    for idx, row in df.iterrows():
        sql = row.get("SqlTextInfo", "")
        if not isinstance(sql, str) or not sql.strip():
            continue

        try:
            pairs = extract_join_pairs(sql, driving_table=str(row.get("Tbl_nm", "")))
        except Exception as exc:  # noqa: BLE001
            logger.warning(
                "Row %s — skipping unparseable SQL. Reason: %s | SQL preview: %.120s",
                idx,
                exc,
                sql,
            )
            skipped += 1
            continue

        if not pairs:
            # No JOIN found — nothing to record
            continue

        for left, right in pairs:
            records.append(
                {
                    "date_wid": row["date_wid"],
                    "Metric date": row["Metric date"],
                    "left_join_table_name": left,
                    "right_join_table_name": right,
                    "user_name": str(row.get("user_name", "")),
                    "app_prefix": row["app_prefix"],
                    "runtime_seconds": row["runtime_seconds"],
                }
            )

    logger.info(
        "Exploded into %d join-pair records. Skipped %d unparseable rows.",
        len(records),
        skipped,
    )
    if not records:
        raise RuntimeError(
            "No join pairs were extracted. Verify that SqlTextInfo contains valid JOIN clauses."
        )

    return pd.DataFrame(records)


def aggregate(exploded: pd.DataFrame) -> pd.DataFrame:
    """Group by date+join-pair and compute all required metrics."""
    group_keys = [
        "date_wid",
        "Metric date",
        "left_join_table_name",
        "right_join_table_name",
    ]

    agg_df = (
        exploded.groupby(group_keys, as_index=False, sort=True)
        .agg(
            join_count=("left_join_table_name", "count"),
            unique_users=("user_name", "nunique"),
            unique_app=("app_prefix", lambda s: s.dropna().nunique()),
            query_count=("runtime_seconds", "count"),
            avg_runtime=("runtime_seconds", "mean"),
        )
    )

    # Round avg_runtime to 2 decimal places (seconds)
    agg_df["avg_runtime"] = agg_df["avg_runtime"].round(2)

    # Post-aggregation sequential Row_id
    agg_df = agg_df.sort_values(
        ["date_wid", "left_join_table_name", "right_join_table_name"]
    ).reset_index(drop=True)
    agg_df.insert(0, "Row_id", range(1, len(agg_df) + 1))

    # Enforce column order
    return agg_df[OUTPUT_COLUMNS]


# ---------------------------------------------------------------------------
# Excel export with professional formatting
# ---------------------------------------------------------------------------

_HEADER_BG   = "1F3864"   # Dark navy blue
_HEADER_FG   = "FFFFFF"   # White text
_ALT_ROW_BG  = "EBF0FA"   # Light blue-grey alternating row
_BORDER_CLR  = "B8C4D6"   # Soft border colour

def _col_widths(df: pd.DataFrame) -> dict[int, float]:
    """Compute sensible column widths based on header + data max length."""
    widths: dict[int, float] = {}
    for i, col in enumerate(df.columns, start=1):
        max_len = max(
            len(str(col)),
            df[col].astype(str).str.len().max() if len(df) > 0 else 0,
        )
        widths[i] = min(max(max_len + 4, 12), 50)
    return widths


def write_excel(result_df: pd.DataFrame, output_path: str) -> None:
    """Write *result_df* to an Excel file with professional styling."""
    logger.info("Writing output to: %s", output_path)

    # Write raw data first via openpyxl
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        result_df.to_excel(writer, index=False, sheet_name="Join_Analysis")

    # Re-open for formatting
    wb = load_workbook(output_path)
    ws = wb["Join_Analysis"]

    thin_border = Border(
        left=Side(style="thin", color=_BORDER_CLR),
        right=Side(style="thin", color=_BORDER_CLR),
        top=Side(style="thin", color=_BORDER_CLR),
        bottom=Side(style="thin", color=_BORDER_CLR),
    )

    # --- Header row ---
    for cell in ws[1]:
        cell.font = Font(bold=True, color=_HEADER_FG, name="Arial", size=10)
        cell.fill = PatternFill("solid", fgColor=_HEADER_BG)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border

    ws.row_dimensions[1].height = 30

    # --- Data rows with alternating colour ---
    alt_fill = PatternFill("solid", fgColor=_ALT_ROW_BG)
    right_cols = {
        i for i, col in enumerate(result_df.columns, start=1)
        if col in ("Row_id", "join_count", "unique_users", "unique_app",
                   "query_count", "avg_runtime", "date_wid")
    }

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
        is_alt = (row_idx % 2 == 0)
        for cell in row:
            cell.font = Font(name="Arial", size=10)
            cell.border = thin_border
            if is_alt:
                cell.fill = alt_fill
            cell.alignment = Alignment(
                horizontal="right" if cell.column in right_cols else "left",
                vertical="center",
            )

    # --- Column widths ---
    widths = _col_widths(result_df)
    for col_idx, width in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # --- Freeze panes below header ---
    ws.freeze_panes = "A2"

    # --- Auto-filter on header ---
    ws.auto_filter.ref = ws.dimensions

    wb.save(output_path)
    logger.info("Output saved successfully: %s  (%d rows)", output_path, len(result_df))


# ---------------------------------------------------------------------------
# CLI entry point
# ---------------------------------------------------------------------------

def build_arg_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        prog="sql_join_parser",
        description=(
            "Parse database query logs from an Excel sheet and produce a "
            "granular join-relationship breakdown in a new Excel sheet."
        ),
        formatter_class=argparse.ArgumentDefaultsHelpFormatter,
    )
    parser.add_argument(
        "input_file",
        help="Path to the input Excel file (.xlsx).",
    )
    parser.add_argument(
        "-o", "--output",
        default=None,
        help=(
            "Path for the output Excel file. "
            "Defaults to <input_stem>_join_analysis.xlsx beside the input file."
        ),
    )
    return parser


def main(argv: list[str] | None = None) -> None:
    parser = build_arg_parser()
    args = parser.parse_args(argv)

    input_path = Path(args.input_file).resolve()
    if not input_path.exists():
        logger.error("Input file not found: %s", input_path)
        sys.exit(1)

    output_path = (
        Path(args.output).resolve()
        if args.output
        else input_path.parent / f"{input_path.stem}_join_analysis.xlsx"
    )

    try:
        raw_df   = load_and_validate_input(str(input_path))
        dated_df = parse_dates(raw_df)
        exploded = explode_join_pairs(dated_df)
        result   = aggregate(exploded)
        write_excel(result, str(output_path))
        logger.info("Done. Total output rows: %d", len(result))
    except Exception as exc:  # noqa: BLE001
        logger.error("Pipeline failed: %s", exc, exc_info=True)
        sys.exit(1)


if __name__ == "__main__":
    main()
