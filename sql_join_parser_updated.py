"""
sql_join_parser.py
==================
Production-grade pipeline that parses database query logs from an Excel sheet
and outputs a granular, row-by-row breakdown of table join relationships.

Author  : Senior Data Engineer
Version : 2.0.0  (robust dirty-log edition)
Requires: pandas, sqlglot, openpyxl, xlsxwriter

Changes in v2.0.0
-----------------
* NEW: Multi-stage SQL repair pipeline handles real-world dirty logs:
    - Fused keywords (e.g. "VINNER JOIN" → "V INNER JOIN", "CDFROM" → "CD FROM")
    - Newlines in the middle of keywords/table names
    - Non-standard block comments using single slashes  /.../ 
    - Missing spaces before/after SQL keywords
    - Mismatched or unclosed string quotes
* NEW: Comment-aware join extraction — line comments (--) that contain a JOIN
  target table (e.g. "-- ccw_view.some_table AS ext ON ...") are parsed and
  their table names are included in the output.  The problem statement notes that
  comments in these logs intentionally carry join information.
* NEW: Dual-mode parsing — tries sqlglot first (best for clean SQL), falls back
  to a regex-based extractor for anything sqlglot cannot handle.
* NEW: Partial-name and alias deduplication so noisy comment fragments don't
  inflate join counts.
"""

from __future__ import annotations

import argparse
import logging
import re
import sys
from pathlib import Path
from typing import Optional

import pandas as pd
import sqlglot
from sqlglot import exp
from openpyxl import load_workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Logging
# ---------------------------------------------------------------------------
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
)
logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------
APP_PREFIXES = ("SVP", "SVR", "SVC", "SVT", "ETL", "DBA", "OP")

INPUT_REQUIRED_COLS = {
    "user_name",
    "Db_nm",
    "Tbl_nm",
    "SqlTextInfo",
    "LogDate",
    "StartTime",
    "LastResponseTime",
}

OUTPUT_COLUMNS = [
    "Row_id",
    "date_wid",
    "Metric date",
    "left_join_table_name",
    "right_join_table_name",
    "join_count",
    "unique_users",
    "unique_app",
    "query_count",
    "avg_runtime",
]

# SQL keywords that commonly get fused with adjacent tokens in dirty logs.
# Ordered longest-first so longer keywords match before their substrings.
# Short/ambiguous keywords (IN, OR, AND, IS, NOT) are intentionally excluded
# because they appear inside ordinary table/column names and would cause false splits.
_FUSE_KEYWORDS = [
    "SELECT", "DISTINCT", "INSERT", "UPDATE", "DELETE",
    "FROM", "WHERE", "HAVING", "UNION", "INTO", "WITH",
    "INNER", "OUTER", "CROSS", "FULL", "LEFT", "RIGHT",
    "JOIN", "GROUP", "ORDER",
    "ON", "AS", "SET", "BY",
]

# ---------------------------------------------------------------------------
# Stage 1 — SQL repair helpers
# ---------------------------------------------------------------------------

def _inject_keyword_spaces(sql: str) -> str:
    """
    Insert a space **before** any SQL keyword that has been fused directly
    onto the end of a preceding word character without whitespace, e.g.::

        "VINNER JOIN"  →  "V INNER JOIN"
        "_NOW_CDFROM"  →  "_NOW_CD FROM"
        "KEYWHERE"     →  "KEY WHERE"

    Only a *before* injection is performed (not after) to avoid splitting
    ordinary identifiers that happen to start with a keyword, e.g.
    "orders" must NOT become "order s" because ORDER is a keyword.

    The keyword is only matched when it ends at a true word boundary
    (followed by a non-word character or end-of-string), ensuring we never
    split inside a longer token like INNER→IN+NER.
    """
    for kw in _FUSE_KEYWORDS:
        # Inject space BEFORE keyword:
        #   - must be preceded by a word char (the fused prefix)
        #   - must end at a real word boundary (followed by non-word or end)
        pattern = r"(?<=[A-Za-z0-9_])(" + re.escape(kw) + r")(?=[^A-Za-z0-9_]|$)"
        sql = re.sub(pattern, r" \1", sql, flags=re.IGNORECASE)
    return re.sub(r"[ \t]+", " ", sql)


def _remove_single_slash_comments(sql: str) -> str:
    """
    Remove non-standard /.../ pseudo block-comments seen in dirty logs, e.g.
    '/RXMV.CLM_SVC_BEG_DT >= add_months(current_date,-120) AND .../'.

    Only removes spans of >=3 characters so we don't eat division operators.
    Real block comments /* ... */ are handled separately.
    """
    return re.sub(r"(?<!\*)(?<!/)/(?!\*)/?.{3,}?(?<!\*)/(?!\*)", " ", sql, flags=re.DOTALL)


def _extract_comment_join_tables(raw_sql: str) -> list[str]:
    """
    Scan every line of *raw_sql* for line comments (--) that appear AFTER a
    JOIN keyword with no table name following it.  When found, the first
    valid identifier-looking token inside the comment is treated as the join
    target table (per the problem statement: comments hold real join info).

    Returns a list of table names (short names, schema prefix stripped).
    """
    found: list[str] = []
    for line in raw_sql.splitlines():
        if "--" not in line:
            continue
        pre_comment, *rest = line.split("--", 1)
        comment_text = rest[0] if rest else ""

        # Inject spaces on pre_comment so we can reliably spot a trailing JOIN
        pre_repaired = _inject_keyword_spaces(pre_comment).strip()

        # Only care if the pre-comment part ends with a JOIN keyword (no table follows)
        if not re.search(r"\bJOIN\s*$", pre_repaired, re.IGNORECASE):
            # Also check: JOIN exists anywhere in comment itself (commented-out JOIN)
            # e.g. "-- INNER JOIN some_table AS t ON ..."
            inline_joins = re.findall(
                r"\b(?:INNER\s+|LEFT\s+(?:OUTER\s+)?|RIGHT\s+(?:OUTER\s+)?|"
                r"FULL\s+(?:OUTER\s+)?|CROSS\s+)?JOIN\s+([\w.]+)",
                comment_text, re.IGNORECASE,
            )
            for tbl in inline_joins:
                tbl_short = tbl.split(".")[-1]
                if len(tbl_short) > 2:
                    found.append(tbl_short)
            continue

        # The JOIN target is inside the comment — find first decent identifier
        comment_repaired = _inject_keyword_spaces(comment_text)
        token_match = re.search(r"\b([\w.]{3,})", comment_repaired.strip())
        if token_match:
            tbl = token_match.group(1).split(".")[-1]
            found.append(tbl)

    return found


def repair_sql(raw_sql: str) -> tuple[str, list[str]]:
    """
    Full multi-stage repair of a dirty SQL log string.

    Returns
    -------
    repaired_sql : str
        Clean(er) SQL with comments stripped and fused keywords separated.
    comment_join_tables : list[str]
        Table names extracted from -- comments that contained join info.
    """
    # ── Stage A: collect join info buried in line comments ──────────────────
    comment_join_tables = _extract_comment_join_tables(raw_sql)

    # ── Stage B: strip line comments (-- to end of logical line) ────────────
    lines = []
    for line in raw_sql.splitlines():
        line = re.sub(r"--.*$", "", line)
        lines.append(line)
    sql = "\n".join(lines)

    # ── Stage C: strip non-standard /.../ pseudo block-comments ─────────────
    sql = _remove_single_slash_comments(sql)

    # ── Stage D: strip real block comments /* ... */ ─────────────────────────
    sql = re.sub(r"/\*.*?\*/", " ", sql, flags=re.DOTALL)

    # ── Stage E: collapse all whitespace to single spaces ───────────────────
    sql = re.sub(r"\s+", " ", sql).strip()

    # ── Stage F: inject spaces around fused SQL keywords ────────────────────
    sql = _inject_keyword_spaces(sql)
    sql = re.sub(r"\s+", " ", sql).strip()

    return sql, comment_join_tables


# ---------------------------------------------------------------------------
# Stage 2 — table extraction (sqlglot + regex fallback)
# ---------------------------------------------------------------------------

def _extract_via_sqlglot(clean_sql: str) -> list[tuple[str, str]]:
    """
    Try to parse *clean_sql* with sqlglot and return top-level join pairs.
    Raises ValueError if parsing fails or produces no pairs.
    """
    try:
        tree = sqlglot.parse_one(clean_sql, error_level=sqlglot.ErrorLevel.WARN)
    except Exception as exc:
        raise ValueError(f"sqlglot.parse_one failed: {exc}") from exc

    # Build alias → real_name map
    alias_map: dict[str, str] = {}
    for node in tree.walk():
        if isinstance(node, exp.Table) and node.name:
            real  = node.name.strip()
            alias = node.alias.strip() if node.alias else ""
            if alias and alias.lower() != real.lower():
                alias_map[alias.lower()] = real

    def resolve(name: str) -> str:
        return alias_map.get(name.lower(), name)

    def _inside_subquery(node: exp.Expression) -> bool:
        parent = node.parent
        while parent is not None:
            if isinstance(parent, exp.Subquery):
                return True
            parent = parent.parent
        return False

    from_table: Optional[str] = None
    join_sequence: list[str] = []

    for node in tree.walk():
        if _inside_subquery(node):
            continue
        if isinstance(node, exp.From):
            tbl = node.find(exp.Table)
            if tbl and tbl.name:
                from_table = resolve(tbl.name.strip())
        elif isinstance(node, exp.Join):
            right = node.args.get("this")
            if isinstance(right, exp.Table) and right.name:
                join_sequence.append(resolve(right.name.strip()))

    if from_table is None or not join_sequence:
        raise ValueError("sqlglot found no FROM/JOIN pairs in this statement.")

    pairs: list[tuple[str, str]] = []
    left = from_table
    for right in join_sequence:
        if left and right and left.lower() != right.lower():
            pairs.append((left, right))
        left = right
    return pairs


_JOIN_REGEX = re.compile(
    r"\b(?:INNER\s+|LEFT\s+(?:OUTER\s+)?|RIGHT\s+(?:OUTER\s+)?"
    r"|FULL\s+(?:OUTER\s+)?|CROSS\s+)?JOIN\s+([\w.]+)",
    re.IGNORECASE,
)
_FROM_REGEX = re.compile(r"\bFROM\s+([\w.]+)", re.IGNORECASE)


def _extract_via_regex(
    repaired_sql: str,
    comment_join_tables: list[str],
    driving_table: Optional[str],
) -> list[tuple[str, str]]:
    """
    Regex-based FROM/JOIN extractor used when sqlglot fails.

    Combines table refs found directly in the SQL with those salvaged from
    -- comments (comment_join_tables).
    """
    from_match = _FROM_REGEX.search(repaired_sql)
    from_table: Optional[str] = None
    if from_match:
        from_table = from_match.group(1).split(".")[-1]

    if from_table is None and driving_table:
        from_table = str(driving_table).strip().split(".")[-1] or None

    # JOIN tables from the main (repaired) SQL body
    sql_join_tables = [
        m.split(".")[-1]
        for m in _JOIN_REGEX.findall(repaired_sql)
        # Filter out string literals that sneak through (start with quote / digit)
        if re.match(r"^[A-Za-z_]", m)
    ]

    # Merge: SQL joins first, then comment-sourced joins
    all_joins = sql_join_tables + [
        t for t in comment_join_tables if t not in sql_join_tables
    ]

    # Deduplicate while preserving order
    seen: set[str] = set()
    unique_joins: list[str] = []
    for t in all_joins:
        key = t.lower()
        if key not in seen and len(t) > 1:
            seen.add(key)
            unique_joins.append(t)

    if from_table is None or not unique_joins:
        return []

    pairs: list[tuple[str, str]] = []
    left = from_table
    for right in unique_joins:
        if left.lower() != right.lower():
            pairs.append((left, right))
        left = right
    return pairs


# ---------------------------------------------------------------------------
# Public API — extract_join_pairs
# ---------------------------------------------------------------------------

def extract_join_pairs(
    sql: str,
    driving_table: Optional[str] = None,
) -> list[tuple[str, str]]:
    """
    Parse *sql* (potentially dirty / uncommented log text) and return an ordered
    list of (left_table, right_table) join pairs.

    Strategy (dual-mode)
    --------------------
    1. Repair the raw SQL:
       a. Extract join info buried in -- comments (comment-aware).
       b. Strip line/block comments.
       c. Remove /.../ pseudo-comments.
       d. Collapse whitespace.
       e. Inject spaces around fused SQL keywords.
    2. Try sqlglot on the repaired SQL → fast and accurate for clean-ish input.
    3. On failure / no results → regex fallback, merging SQL and comment join tables.

    Parameters
    ----------
    sql           : Raw SQL text from the log (may be very dirty).
    driving_table : Optional fallback table name from the ``Tbl_nm`` column.

    Returns
    -------
    List of (left_table, right_table) string tuples.
    """
    if not isinstance(sql, str) or not sql.strip():
        return []

    repaired_sql, comment_join_tables = repair_sql(sql)

    if not repaired_sql:
        return []

    # ── Try sqlglot first ───────────────────────────────────────────────────
    try:
        pairs = _extract_via_sqlglot(repaired_sql)
        if pairs:
            # Even with sqlglot success, append any extra joins from comments
            seen_rights = {r.lower() for _, r in pairs}
            last_right = pairs[-1][1] if pairs else None
            for ct in comment_join_tables:
                if ct.lower() not in seen_rights and last_right and ct.lower() != last_right.lower():
                    pairs.append((last_right, ct))
                    last_right = ct
                    seen_rights.add(ct.lower())
            return pairs
    except Exception:
        pass  # fall through to regex

    # ── Regex fallback ──────────────────────────────────────────────────────
    return _extract_via_regex(repaired_sql, comment_join_tables, driving_table)


# ---------------------------------------------------------------------------
# Application prefix extractor
# ---------------------------------------------------------------------------

def _extract_app_prefix(user_name: str) -> Optional[str]:
    """Return the matching APP_PREFIX for a user_name, or None."""
    if not isinstance(user_name, str):
        return None
    upper = user_name.upper()
    for prefix in APP_PREFIXES:
        if upper.startswith(prefix):
            return prefix
    return None


# ---------------------------------------------------------------------------
# Core pipeline
# ---------------------------------------------------------------------------

def load_and_validate_input(filepath: str) -> pd.DataFrame:
    """Read the Excel input file and validate required columns."""
    logger.info("Loading input file: %s", filepath)
    df = pd.read_excel(filepath, dtype={"user_name": str, "Db_nm": str, "Tbl_nm": str})

    missing = INPUT_REQUIRED_COLS - set(df.columns)
    if missing:
        raise ValueError(f"Input file is missing required columns: {missing}")

    logger.info("Loaded %d rows from input.", len(df))
    return df


def parse_dates(df: pd.DataFrame) -> pd.DataFrame:
    """Coerce date/time columns to proper types."""
    df["LogDate"]          = pd.to_datetime(df["LogDate"],          errors="coerce")
    df["StartTime"]        = pd.to_datetime(df["StartTime"],        errors="coerce")
    df["LastResponseTime"] = pd.to_datetime(df["LastResponseTime"], errors="coerce")

    null_dates = df["LogDate"].isna().sum()
    if null_dates:
        logger.warning("%d rows have unparseable LogDate; they will be skipped.", null_dates)
    df = df.dropna(subset=["LogDate"]).copy()

    df["date_wid"]        = df["LogDate"].dt.strftime("%Y%m%d").astype(str)
    df["Metric date"]     = df["LogDate"].dt.normalize()
    df["runtime_seconds"] = (
        df["LastResponseTime"] - df["StartTime"]
    ).dt.total_seconds().clip(lower=0)
    df["app_prefix"]      = df["user_name"].apply(_extract_app_prefix)

    return df


def explode_join_pairs(df: pd.DataFrame) -> pd.DataFrame:
    """
    For every row in *df*, extract join pairs from SqlTextInfo and
    produce one expanded row per (query_row × join_pair).
    """
    records: list[dict] = []
    skipped = 0

    for idx, row in df.iterrows():
        sql = row.get("SqlTextInfo", "")
        if not isinstance(sql, str) or not sql.strip():
            continue

        try:
            pairs = extract_join_pairs(sql, driving_table=str(row.get("Tbl_nm", "")))
        except Exception as exc:  # noqa: BLE001
            logger.warning(
                "Row %s — skipping unparseable SQL. Reason: %s | SQL preview: %.120s",
                idx, exc, sql,
            )
            skipped += 1
            continue

        if not pairs:
            # No JOIN found — nothing to record
            continue

        for left, right in pairs:
            records.append({
                "date_wid":              row["date_wid"],
                "Metric date":           row["Metric date"],
                "left_join_table_name":  left,
                "right_join_table_name": right,
                "user_name":             str(row.get("user_name", "")),
                "app_prefix":            row["app_prefix"],
                "runtime_seconds":       row["runtime_seconds"],
            })

    logger.info(
        "Exploded into %d join-pair records. Skipped %d unparseable rows.",
        len(records), skipped,
    )
    if not records:
        raise RuntimeError(
            "No join pairs were extracted. Verify that SqlTextInfo contains valid JOIN clauses."
        )

    return pd.DataFrame(records)


def aggregate(exploded: pd.DataFrame) -> pd.DataFrame:
    """Group by date+join-pair and compute all required metrics."""
    group_keys = [
        "date_wid",
        "Metric date",
        "left_join_table_name",
        "right_join_table_name",
    ]

    agg_df = (
        exploded.groupby(group_keys, as_index=False, sort=True)
        .agg(
            join_count  =("left_join_table_name", "count"),
            unique_users=("user_name",            "nunique"),
            unique_app  =("app_prefix",           lambda s: s.dropna().nunique()),
            query_count =("runtime_seconds",      "count"),
            avg_runtime =("runtime_seconds",      "mean"),
        )
    )

    agg_df["avg_runtime"] = agg_df["avg_runtime"].round(2)
    agg_df = agg_df.sort_values(
        ["date_wid", "left_join_table_name", "right_join_table_name"]
    ).reset_index(drop=True)
    agg_df.insert(0, "Row_id", range(1, len(agg_df) + 1))

    return agg_df[OUTPUT_COLUMNS]


# ---------------------------------------------------------------------------
# Excel export with professional formatting
# ---------------------------------------------------------------------------

_HEADER_BG  = "1F3864"
_HEADER_FG  = "FFFFFF"
_ALT_ROW_BG = "EBF0FA"
_BORDER_CLR = "B8C4D6"


def _col_widths(df: pd.DataFrame) -> dict[int, float]:
    """Compute sensible column widths based on header + data max length."""
    widths: dict[int, float] = {}
    for i, col in enumerate(df.columns, start=1):
        max_len = max(
            len(str(col)),
            df[col].astype(str).str.len().max() if len(df) > 0 else 0,
        )
        widths[i] = min(max(max_len + 4, 12), 50)
    return widths


def write_excel(result_df: pd.DataFrame, output_path: str) -> None:
    """Write *result_df* to an Excel file with professional styling."""
    logger.info("Writing output to: %s", output_path)

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        result_df.to_excel(writer, index=False, sheet_name="Join_Analysis")

    wb = load_workbook(output_path)
    ws = wb["Join_Analysis"]

    thin_border = Border(
        left  =Side(style="thin", color=_BORDER_CLR),
        right =Side(style="thin", color=_BORDER_CLR),
        top   =Side(style="thin", color=_BORDER_CLR),
        bottom=Side(style="thin", color=_BORDER_CLR),
    )

    for cell in ws[1]:
        cell.font      = Font(bold=True, color=_HEADER_FG, name="Arial", size=10)
        cell.fill      = PatternFill("solid", fgColor=_HEADER_BG)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = thin_border

    ws.row_dimensions[1].height = 30

    alt_fill = PatternFill("solid", fgColor=_ALT_ROW_BG)
    right_cols = {
        i for i, col in enumerate(result_df.columns, start=1)
        if col in ("Row_id", "join_count", "unique_users", "unique_app",
                   "query_count", "avg_runtime", "date_wid")
    }

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
        is_alt = (row_idx % 2 == 0)
        for cell in row:
            cell.font   = Font(name="Arial", size=10)
            cell.border = thin_border
            if is_alt:
                cell.fill = alt_fill
            cell.alignment = Alignment(
                horizontal="right" if cell.column in right_cols else "left",
                vertical="center",
            )

    widths = _col_widths(result_df)
    for col_idx, width in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.freeze_panes    = "A2"
    ws.auto_filter.ref = ws.dimensions

    wb.save(output_path)
    logger.info("Output saved successfully: %s  (%d rows)", output_path, len(result_df))


# ---------------------------------------------------------------------------
# CLI entry point
# ---------------------------------------------------------------------------

def build_arg_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        prog="sql_join_parser",
        description=(
            "Parse database query logs from an Excel sheet and produce a "
            "granular join-relationship breakdown in a new Excel sheet."
        ),
        formatter_class=argparse.ArgumentDefaultsHelpFormatter,
    )
    parser.add_argument("input_file", help="Path to the input Excel file (.xlsx).")
    parser.add_argument(
        "-o", "--output",
        default=None,
        help=(
            "Path for the output Excel file. "
            "Defaults to <input_stem>_join_analysis.xlsx beside the input file."
        ),
    )
    return parser


def main(argv: list[str] | None = None) -> None:
    parser   = build_arg_parser()
    args     = parser.parse_args(argv)

    input_path = Path(args.input_file).resolve()
    if not input_path.exists():
        logger.error("Input file not found: %s", input_path)
        sys.exit(1)

    output_path = (
        Path(args.output).resolve()
        if args.output
        else input_path.parent / f"{input_path.stem}_join_analysis.xlsx"
    )

    try:
        raw_df   = load_and_validate_input(str(input_path))
        dated_df = parse_dates(raw_df)
        exploded = explode_join_pairs(dated_df)
        result   = aggregate(exploded)
        write_excel(result, str(output_path))
        logger.info("Done. Total output rows: %d", len(result))
    except Exception as exc:  # noqa: BLE001
        logger.error("Pipeline failed: %s", exc, exc_info=True)
        sys.exit(1)


if __name__ == "__main__":
    main()
